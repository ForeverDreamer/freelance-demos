"""Excel writer producing client-format Buyer Database xlsx.

Output layout matches the client's `Buyers_Database_Final3.xlsx` Buyer Database sheet:
- Row 1: section headings (FIRM IDENTITY / INVESTMENT OVERVIEW / ...) merged across each section
- Row 2: column names (33 cols)
- Row 3+: data rows (one per firm)

Multi-value fields (key_contacts, lists) serialized as readable bullet strings
(not JSON), since the client opens this in Excel manually.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import EXCEL_COLUMN_LABELS, EXCEL_SECTIONS, FirmRecord, KeyContact

# Visual style aligned with client template (see Buyers_Database_Final3.xlsx)
SECTION_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="000000")
DATA_FONT = Font(name="Calibri", size=10)
EXTERNAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _format_key_contacts(contacts: list[KeyContact]) -> str:
    if not contacts:
        return ""
    parts = []
    for c in contacts:
        bits = []
        if c.name:
            bits.append(c.name)
        if c.title:
            bits.append(c.title)
        if c.email:
            bits.append(c.email)
        parts.append(" | ".join(bits))
    return "\n".join(parts)


def _format_value(field_name: str, value) -> str:
    """Convert a record field value to a human-readable Excel cell string."""
    if value is None:
        return ""
    if field_name == "key_contacts":
        return _format_key_contacts(value)
    if field_name == "source_urls":
        return "\n".join(str(u) for u in value)
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if hasattr(value, "value"):  # Enum
        return value.value
    return str(value)


def write_buyer_database(records: list[FirmRecord], output_path: Path) -> None:
    """Write a client-format xlsx with proper section headers and 33 columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyer Database"

    field_order = [field for _, fields in EXCEL_SECTIONS for field in fields]
    if len(field_order) != len(EXCEL_COLUMN_LABELS):
        raise RuntimeError(
            f"Schema mismatch: {len(field_order)} fields in EXCEL_SECTIONS vs "
            f"{len(EXCEL_COLUMN_LABELS)} in EXCEL_COLUMN_LABELS"
        )

    # Row 1: section headings (merged across each section's column span)
    col_idx = 1
    for section_name, section_fields in EXCEL_SECTIONS:
        first_col = col_idx
        last_col = col_idx + len(section_fields) - 1
        cell = ws.cell(row=1, column=first_col, value=section_name)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if last_col > first_col:
            ws.merge_cells(
                start_row=1, start_column=first_col, end_row=1, end_column=last_col
            )
            for c in range(first_col + 1, last_col + 1):
                fill_cell = ws.cell(row=1, column=c)
                fill_cell.fill = SECTION_FILL
                fill_cell.border = THIN_BORDER
        col_idx = last_col + 1

    # Row 2: column labels
    for i, field_name in enumerate(field_order, start=1):
        cell = ws.cell(row=2, column=i, value=EXCEL_COLUMN_LABELS[field_name])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Row 3+: data
    for row_offset, record in enumerate(records, start=3):
        record_dict = record.model_dump()
        for i, field_name in enumerate(field_order, start=1):
            raw = record_dict.get(field_name)
            value = _format_value(field_name, getattr(record, field_name))
            cell = ws.cell(row=row_offset, column=i, value=value)
            cell.font = DATA_FONT
            cell.alignment = LEFT_WRAP
            cell.border = THIN_BORDER
            if isinstance(raw, str) and raw.startswith("[REQUIRES EXTERNAL"):
                cell.fill = EXTERNAL_FILL

    # Column widths (rough heuristic; firm_name and notes wider)
    width_overrides = {
        "firm_name": 22,
        "website": 30,
        "hq_location": 20,
        "key_contacts": 35,
        "platform_industries": 30,
        "addon_industries": 30,
        "platform_geographies": 20,
        "addon_geographies": 20,
        "transaction_types": 25,
        "platform_additional_criteria": 30,
        "addon_additional_criteria": 30,
        "addon_opportunistic": 25,
        "source_urls": 35,
        "notes": 40,
    }
    for i, field_name in enumerate(field_order, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width_overrides.get(
            field_name, 18
        )

    # Freeze panes below header
    ws.freeze_panes = "A3"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
